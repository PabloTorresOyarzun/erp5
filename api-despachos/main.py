from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File, Query, Form
from pydantic import BaseModel
from sqlalchemy import create_engine, Column, String, Text, DateTime, Boolean, Integer, JSON
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
import os
import requests
import json
from datetime import datetime
from typing import List, Optional, Dict, Any
import base64

app = FastAPI()

# Configuración de base de datos
DATABASE_URL = f"postgresql://postgres:postgres@postgres/postgres"
SGD_URL = os.getenv('SGD_URL')
SGD_AUTH_TOKEN = os.getenv('SGD_AUTH_TOKEN')

# SQLAlchemy
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# Modelos de base de datos
class Despacho(Base):
    __tablename__ = "despachos"
    __table_args__ = {"schema": "operaciones"} 
    
    numero_despacho = Column(String, primary_key=True)
    estado = Column(String, default="pendiente")
    fecha_creacion = Column(DateTime, default=datetime.now)
    fecha_actualizacion = Column(DateTime, default=datetime.now, onupdate=datetime.now)
    documentos_requeridos = Column(JSON)
    documentos_presentes = Column(JSON)
    datos_extraidos = Column(JSON)
    extra_metadata = Column(JSON)

class Documento(Base):
    __tablename__ = "documentos"
    __table_args__ = {"schema": "operaciones"}
    
    id = Column(Integer, primary_key=True, index=True)
    numero_despacho = Column(String, index=True)
    tipo_documento = Column(String)
    nombre_archivo = Column(String)
    contenido_base64 = Column(Text)
    procesado = Column(Boolean, default=False)
    datos_extraidos = Column(JSON)
    fecha_carga = Column(DateTime, default=datetime.now)
    fecha_procesamiento = Column(DateTime, nullable=True)

class Procedimiento(Base):
    __tablename__ = "procedimientos"
    __table_args__ = {"schema": "operaciones"}
    
    id = Column(Integer, primary_key=True, index=True)
    numero_despacho = Column(String, index=True)
    tipo_procedimiento = Column(String)
    estado = Column(String, default="pendiente")
    usuario_asignado = Column(String, nullable=True)
    fecha_inicio = Column(DateTime, nullable=True)
    fecha_fin = Column(DateTime, nullable=True)
    datos = Column(JSON)

# Crear tablas
Base.metadata.create_all(bind=engine)

# Dependencia para obtener la sesión de BD
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Modelos Pydantic
class DespachoCreate(BaseModel):
    numero_despacho: str
    extra_metadata: Optional[Dict[str, Any]] = {}

class DocumentoUpload(BaseModel):
    numero_despacho: str
    tipo_documento: str
    nombre_archivo: str
    contenido_base64: str

class DespachoStatus(BaseModel):
    numero_despacho: str
    estado: str
    documentos_requeridos: List[str]
    documentos_presentes: List[str]
    documentos_faltantes: List[str]
    porcentaje_completitud: float
    puede_procesar: bool
    procedimientos: List[Dict[str, Any]]

class ProcedimientoCreate(BaseModel):
    numero_despacho: str
    tipo_procedimiento: str
    usuario_asignado: Optional[str] = None

# Configuración de tipos de documentos requeridos FIJOS
DOCUMENTOS_REQUERIDOS = [
    "factura_comercial",
    "documento_transporte"
]

# Endpoints
@app.get("/")
async def root():
    return {"service": "Despachos Management API", "status": "running"}

@app.get("/health")
async def health():
    return {
        "status": "healthy",
        "service": "api-despachos",
        "database": "connected",
        "sgd_configured": bool(SGD_URL and SGD_AUTH_TOKEN)
    }

@app.post("/despachos/crear")
async def crear_despacho(
    despacho_data: DespachoCreate,
    db: Session = Depends(get_db)
):
    """Crear un nuevo despacho"""
    existing = db.query(Despacho).filter(
        Despacho.numero_despacho == despacho_data.numero_despacho
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="El despacho ya existe")
    
    nuevo_despacho = Despacho(
        numero_despacho=despacho_data.numero_despacho,
        documentos_requeridos=DOCUMENTOS_REQUERIDOS,  # Siempre los mismos
        documentos_presentes=[],
        datos_extraidos={},
        extra_metadata=despacho_data.extra_metadata
    )
    
    db.add(nuevo_despacho)
    db.commit()
    db.refresh(nuevo_despacho)
    
    return {"message": "Despacho creado exitosamente", "numero_despacho": nuevo_despacho.numero_despacho}

@app.get("/despachos/{numero_despacho}/sgd")
async def obtener_documentos_sgd(
    numero_despacho: str,
    db: Session = Depends(get_db)
):
    """Obtener documentos desde SGD"""
    if not SGD_URL or not SGD_AUTH_TOKEN:
        raise HTTPException(status_code=500, detail="SGD no configurado")
    
    # Verificar/crear despacho
    despacho = db.query(Despacho).filter(
        Despacho.numero_despacho == numero_despacho
    ).first()
    
    if not despacho:
        despacho = Despacho(
            numero_despacho=numero_despacho,
            documentos_requeridos=DOCUMENTOS_REQUERIDOS,
            documentos_presentes=[],
            datos_extraidos={},
            extra_metadata={}
        )
        db.add(despacho)
        db.commit()
    
    try:
        # Construir URL
        url = f"{SGD_URL.rstrip('/')}/{numero_despacho}"
        
        print(f"\n🔍 CONSULTANDO SGD:")
        print(f"   URL: {url}")
        print(f"   Token: Bearer {SGD_AUTH_TOKEN[:20]}...")
        
        headers = {
            "Authorization": f"Bearer {SGD_AUTH_TOKEN}",
            "Accept": "application/json"
        }
        
        # Hacer la petición
        response = requests.get(url, headers=headers, timeout=30)
        
        print(f"📥 RESPUESTA:")
        print(f"   Status: {response.status_code}")
        
        if response.status_code == 404:
            return {
                "message": "Despacho no encontrado en SGD",
                "total": 0
            }
        
        response.raise_for_status()
        
        # Parsear JSON
        json_response = response.json()
        
        # La respuesta debe tener una estructura como {"data": [...]}
        documentos = json_response.get('data', [])
        
        if not isinstance(documentos, list):
            print(f"⚠️ Respuesta no es una lista: {type(documentos)}")
            return {
                "message": "Formato de respuesta inesperado",
                "total": 0
            }
        
        print(f"📄 Documentos encontrados: {len(documentos)}")
        
        documentos_importados = []
        
        for idx, doc in enumerate(documentos):
            try:
                # Extraer el base64
                documento_base64 = doc.get('documento', '')
                
                # Si viene con prefijo data:, quitarlo
                if documento_base64.startswith("data:application/pdf;base64,"):
                    documento_base64 = documento_base64.split(",", 1)[1]
                
                # Obtener nombre
                nombre = doc.get('nombre_documento', f'documento_{idx+1}.pdf')
                
                # Determinar tipo por nombre
                tipo = 'general'
                nombre_lower = nombre.lower()
                
                # Mejorar detección - ACTUALIZADO
                if 'factura' in nombre_lower or 'invoice' in nombre_lower or 'facture' in nombre_lower:
                    tipo = 'factura_comercial'
                elif any(term in nombre_lower for term in ['bl', 'awb', 'waybill', 'transporte', 'transport', 'bill-of-lading', 'embarque']):
                    tipo = 'documento_transporte'
                elif 'packing' in nombre_lower or 'lista-empaque' in nombre_lower:
                    tipo = 'packing_list'
                elif 'certificado' in nombre_lower or 'certificate' in nombre_lower or 'origen' in nombre_lower:
                    tipo = 'certificado_origen'
                
                print(f"   - Documento {idx+1}: {nombre} ({tipo})")
                
                # Guardar en BD
                nuevo_doc = Documento(
                    numero_despacho=numero_despacho,
                    tipo_documento=tipo,
                    nombre_archivo=nombre,
                    contenido_base64=documento_base64,
                    procesado=False
                )
                db.add(nuevo_doc)
                documentos_importados.append(tipo)
                
            except Exception as e:
                print(f"   ❌ Error procesando documento {idx}: {e}")
                continue
        
        # Actualizar despacho
        if documentos_importados:
            documentos_actuales = despacho.documentos_presentes or []
            documentos_actuales.extend(documentos_importados)
            despacho.documentos_presentes = list(set(documentos_actuales))
            db.commit()
        
        print(f"✅ Importados: {len(documentos_importados)} documentos")
        
        return {
            "message": "Documentos importados desde SGD",
            "documentos_importados": documentos_importados,
            "total": len(documentos_importados)
        }
        
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            return {
                "message": "Despacho no encontrado en SGD",
                "total": 0
            }
        elif e.response.status_code == 401:
            raise HTTPException(status_code=401, detail="No autorizado en SGD")
        else:
            raise HTTPException(
                status_code=e.response.status_code,
                detail=f"Error SGD: {e.response.text}"
            )
    except requests.exceptions.RequestException as e:
        print(f"❌ Error de conexión: {e}")
        raise HTTPException(status_code=500, detail=f"Error conectando con SGD: {str(e)}")
    except Exception as e:
        print(f"❌ Error general: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando respuesta: {str(e)}")

# Modificar endpoint en api-despachos/main.py

@app.post("/despachos/{numero_despacho}/documento/subir")
async def subir_documento(
    numero_despacho: str,
    tipo_documento: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Subir un documento a un despacho con procesamiento automático opcional"""
    
    # Verificar despacho existe
    despacho = db.query(Despacho).filter(Despacho.numero_despacho == numero_despacho).first()
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    try:
        contents = await file.read()
        
        if not contents.startswith(b'%PDF'):
            raise HTTPException(status_code=400, detail="El archivo no es un PDF válido")
        
        # Si el tipo es automático, enviar a API-DOCS para procesamiento
        if tipo_documento == "automatico":
            import requests
            
            # Preparar archivo para envío
            files = {'file': (file.filename, contents, 'application/pdf')}
            data = {'numero_despacho': numero_despacho}
            
            # Llamar a API-DOCS para procesamiento automático
            response = requests.post(
                f"{os.getenv('DOC_API_URL', 'http://api-docs:8002')}/process/automatic",
                files=files,
                data=data
            )
            
            if response.status_code == 200:
                resultado = response.json()
                
                # Guardar cada documento identificado
                documentos_guardados = []
                for doc in resultado['resultado']['documentos']:
                    nuevo_doc = Documento(
                        numero_despacho=numero_despacho,
                        tipo_documento=doc['tipo'],
                        nombre_archivo=f"{doc['id']}.pdf",
                        contenido_base64=doc['pdf_base64'],
                        datos_extraidos=json.dumps(doc['datos_extraidos']),
                        procesado=doc['procesado']
                    )
                    db.add(nuevo_doc)
                    documentos_guardados.append({
                        "id": doc['id'],
                        "tipo": doc['tipo'],
                        "paginas": doc['paginas']
                    })
                
                db.commit()
                
                return {
                    "message": "Documentos procesados e identificados automáticamente",
                    "total_documentos": len(documentos_guardados),
                    "documentos": documentos_guardados,
                    "proceso_id": resultado['id']
                }
            else:
                raise HTTPException(
                    status_code=500,
                    detail=f"Error en procesamiento automático: {response.text}"
                )
        
        # Si no es automático, guardar como antes
        else:
            contenido_b64 = base64.b64encode(contents).decode('utf-8')
            
            nuevo_documento = Documento(
                numero_despacho=numero_despacho,
                tipo_documento=tipo_documento,
                nombre_archivo=file.filename,
                contenido_base64=contenido_b64,
                procesado=False
            )
            
            db.add(nuevo_documento)
            
            # Actualizar documentos presentes
            documentos_actuales = despacho.documentos_presentes or []
            if tipo_documento not in documentos_actuales:
                documentos_actuales.append(tipo_documento)
                despacho.documentos_presentes = documentos_actuales
            
            db.commit()
            
            return {
                "message": "Documento subido exitosamente",
                "documento_id": nuevo_documento.id,
                "tipo": tipo_documento
            }
            
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/despachos/{numero_despacho}/exportar/json")
async def exportar_json(numero_despacho: str, db: Session = Depends(get_db)):
    """Exportar todos los datos procesados del despacho como JSON"""
    
    despacho = db.query(Despacho).filter(Despacho.numero_despacho == numero_despacho).first()
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    # Obtener todos los documentos
    documentos = db.query(Documento).filter(
        Documento.numero_despacho == numero_despacho
    ).all()
    
    # Construir respuesta estructurada
    resultado = {
        "numero_despacho": numero_despacho,
        "estado": despacho.estado,
        "fecha_creacion": despacho.fecha_creacion.isoformat(),
        "documentos": []
    }
    
    for doc in documentos:
        doc_data = {
            "id": doc.id,
            "tipo": doc.tipo_documento,
            "nombre": doc.nombre_archivo,
            "procesado": doc.procesado,
            "fecha_subida": doc.fecha_subida.isoformat()
        }
        
        # Incluir datos extraídos si existen
        if doc.datos_extraidos:
            try:
                doc_data["datos_extraidos"] = json.loads(doc.datos_extraidos)
            except:
                doc_data["datos_extraidos"] = doc.datos_extraidos
        
        resultado["documentos"].append(doc_data)
    
    return resultado

@app.get("/despachos/{numero_despacho}/exportar/excel")
async def exportar_excel(numero_despacho: str, db: Session = Depends(get_db)):
    """Exportar datos del despacho como Excel para Declaración de Ingreso"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import io
    
    despacho = db.query(Despacho).filter(Despacho.numero_despacho == numero_despacho).first()
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    documentos = db.query(Documento).filter(
        Documento.numero_despacho == numero_despacho
    ).all()
    
    # Crear libro Excel
    wb = Workbook()
    
    # Hoja 1: Declaración de Ingreso (formato aduanas Chile)
    ws_din = wb.active
    ws_din.title = "Declaración de Ingreso"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezado principal
    ws_din['A1'] = "DECLARACIÓN DE INGRESO - ADUANA CHILE"
    ws_din['A1'].font = Font(bold=True, size=16)
    ws_din.merge_cells('A1:H1')
    
    ws_din['A3'] = "Número Despacho:"
    ws_din['B3'] = numero_despacho
    
    # Sección: Datos del Importador/Consignatario
    ws_din['A5'] = "DATOS DEL IMPORTADOR"
    ws_din['A5'].font = Font(bold=True, size=12)
    ws_din['A5'].fill = header_fill
    ws_din['A5'].font = header_font
    ws_din.merge_cells('A5:D5')
    
    # Buscar datos del consignatario desde documentos procesados
    row = 6
    for doc in documentos:
        if doc.datos_extraidos and doc.tipo_documento in ['factura', 'transporte']:
            try:
                datos = json.loads(doc.datos_extraidos)
                
                # Extraer info del cliente/consignatario
                if 'customer_information' in datos:
                    for campo, valor in datos['customer_information'].items():
                        ws_din[f'A{row}'] = campo.replace('_', ' ').title()
                        ws_din[f'B{row}'] = valor
                        row += 1
                elif 'consignee' in datos:
                    for campo, valor in datos['consignee'].items():
                        ws_din[f'A{row}'] = campo.replace('_', ' ').title()
                        ws_din[f'B{row}'] = valor
                        row += 1
                break
            except:
                pass
    
    row += 1
    
    # Sección: Datos de la Mercancía
    ws_din[f'A{row}'] = "DATOS DE LA MERCANCÍA"
    ws_din[f'A{row}'].font = Font(bold=True, size=12)
    ws_din[f'A{row}'].fill = header_fill
    ws_din[f'A{row}'].font = header_font
    ws_din.merge_cells(f'A{row}:H{row}')
    row += 1
    
    # Encabezados de tabla
    headers = ['Item', 'Descripción', 'Cantidad', 'Unidad', 'Valor Unit.', 'Valor Total', 'Peso Neto', 'Peso Bruto']
    for col, header in enumerate(headers, 1):
        cell = ws_din.cell(row=row, column=col)
        cell.value = header
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
    row += 1
    
    # Buscar items desde facturas
    item_num = 1
    for doc in documentos:
        if doc.datos_extraidos and doc.tipo_documento == 'factura':
            try:
                datos = json.loads(doc.datos_extraidos)
                if 'line_items' in datos:
                    for item in datos['line_items']:
                        ws_din[f'A{row}'] = item_num
                        # Mapear campos del item a columnas
                        for col, valor in enumerate(item.values(), 2):
                            ws_din.cell(row=row, column=col).value = valor
                        row += 1
                        item_num += 1
            except:
                pass
    
    # Hoja 2: Resumen de Documentos
    ws_docs = wb.create_sheet("Documentos")
    ws_docs['A1'] = "RESUMEN DE DOCUMENTOS"
    ws_docs['A1'].font = Font(bold=True, size=14)
    
    headers = ['Tipo', 'Nombre Archivo', 'Procesado', 'Fecha']
    for col, header in enumerate(headers, 1):
        cell = ws_docs.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    row = 4
    for doc in documentos:
        ws_docs[f'A{row}'] = doc.tipo_documento
        ws_docs[f'B{row}'] = doc.nombre_archivo
        ws_docs[f'C{row}'] = "Sí" if doc.procesado else "No"
        ws_docs[f'D{row}'] = doc.fecha_subida.strftime('%Y-%m-%d %H:%M')
        row += 1
    
    # Ajustar anchos de columna
    for ws in [ws_din, ws_docs]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    from fastapi.responses import StreamingResponse
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=DIN_{numero_despacho}.xlsx"
        }
    )

@app.get("/despachos/{numero_despacho}/estado")
async def obtener_estado_despacho(
    numero_despacho: str,
    db: Session = Depends(get_db)
):
    """Obtener el estado completo de un despacho"""
    despacho = db.query(Despacho).filter(
        Despacho.numero_despacho == numero_despacho
    ).first()
    
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    documentos_requeridos = DOCUMENTOS_REQUERIDOS  # Siempre los mismos
    documentos_presentes = despacho.documentos_presentes or []
    documentos_faltantes = [doc for doc in documentos_requeridos if doc not in documentos_presentes]
    
    porcentaje_completitud = (len(documentos_presentes) / len(documentos_requeridos) * 100) if documentos_requeridos else 0
    
    puede_procesar = len(documentos_faltantes) == 0
    
    procedimientos = db.query(Procedimiento).filter(
        Procedimiento.numero_despacho == numero_despacho
    ).all()
    
    procedimientos_data = []
    for proc in procedimientos:
        procedimientos_data.append({
            "id": proc.id,
            "tipo": proc.tipo_procedimiento,
            "estado": proc.estado,
            "usuario_asignado": proc.usuario_asignado,
            "fecha_inicio": proc.fecha_inicio.isoformat() if proc.fecha_inicio else None,
            "fecha_fin": proc.fecha_fin.isoformat() if proc.fecha_fin else None
        })
    
    return DespachoStatus(
        numero_despacho=numero_despacho,
        estado=despacho.estado,
        documentos_requeridos=documentos_requeridos,
        documentos_presentes=documentos_presentes,
        documentos_faltantes=documentos_faltantes,
        porcentaje_completitud=porcentaje_completitud,
        puede_procesar=puede_procesar,
        procedimientos=procedimientos_data
    )

@app.get("/despachos/{numero_despacho}/documentos")
async def listar_documentos_despacho(
    numero_despacho: str,
    db: Session = Depends(get_db)
):
    """Listar todos los documentos de un despacho"""
    documentos = db.query(Documento).filter(
        Documento.numero_despacho == numero_despacho
    ).all()
    
    documentos_data = []
    for doc in documentos:
        documentos_data.append({
            "id": doc.id,
            "tipo_documento": doc.tipo_documento,
            "nombre_archivo": doc.nombre_archivo,
            "procesado": doc.procesado,
            "fecha_carga": doc.fecha_carga.isoformat(),
            "fecha_procesamiento": doc.fecha_procesamiento.isoformat() if doc.fecha_procesamiento else None,
            "tiene_contenido": bool(doc.contenido_base64),
            "tiene_datos": bool(doc.datos_extraidos)
        })
    
    return documentos_data

@app.get("/despachos/{numero_despacho}/documento/{documento_id}/pdf")
async def obtener_documento_pdf(
    numero_despacho: str,
    documento_id: int,
    db: Session = Depends(get_db)
):
    """Obtener un documento en formato PDF"""
    documento = db.query(Documento).filter(
        Documento.id == documento_id,
        Documento.numero_despacho == numero_despacho
    ).first()
    
    if not documento:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    if not documento.contenido_base64:
        raise HTTPException(status_code=404, detail="El documento no tiene contenido")
    
    try:
        pdf_content = base64.b64decode(documento.contenido_base64)
        
        from fastapi.responses import Response
        return Response(
            content=pdf_content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"inline; filename={documento.nombre_archivo}"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error decodificando PDF: {str(e)}")

@app.get("/despachos/{numero_despacho}/documento/{documento_id}/json")
async def obtener_documento_json(
    numero_despacho: str,
    documento_id: int,
    db: Session = Depends(get_db)
):
    """Obtener datos extraídos del documento en formato JSON"""
    documento = db.query(Documento).filter(
        Documento.id == documento_id,
        Documento.numero_despacho == numero_despacho
    ).first()
    
    if not documento:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    if not documento.datos_extraidos:
        raise HTTPException(status_code=404, detail="El documento no tiene datos procesados")
    
    from fastapi.responses import JSONResponse
    return JSONResponse(
        content=documento.datos_extraidos,
        headers={
            "Content-Disposition": f"attachment; filename={documento.nombre_archivo}.json"
        }
    )

@app.post("/despachos/{numero_despacho}/documento/{documento_id}/procesar")
async def procesar_documento_individual(
    numero_despacho: str,
    documento_id: int,
    db: Session = Depends(get_db),
    authorization: str = Header(None)
):
    """Procesar un documento individual"""
    documento = db.query(Documento).filter(
        Documento.id == documento_id,
        Documento.numero_despacho == numero_despacho
    ).first()
    
    if not documento:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    if not documento.contenido_base64:
        raise HTTPException(status_code=400, detail="El documento no tiene contenido")
    
    try:
        pdf_bytes = base64.b64decode(documento.contenido_base64)
        
        # Determinar endpoint según tipo
        endpoint = "invoice" if "factura" in documento.tipo_documento.lower() else "transport"
        
        from io import BytesIO
        files = {
            'file': (documento.nombre_archivo, BytesIO(pdf_bytes), 'application/pdf')
        }
        
        # Agregar número de despacho y tipo para guardar en BD
        data = {
            'numero_despacho': numero_despacho,
            'tipo_documento': documento.tipo_documento
        }
        
        headers = {}
        if authorization:
            headers['Authorization'] = authorization
        
        response = requests.post(
            f"http://api-docs:8002/process/{endpoint}",
            files=files,
            data=data,
            headers=headers
        )
        
        if response.status_code == 200:
            result = response.json()
            
            # Actualizar documento
            documento.datos_extraidos = result.get('extracted_data', {})
            documento.procesado = True
            documento.fecha_procesamiento = datetime.now()
            
            # Actualizar despacho
            despacho = db.query(Despacho).filter(
                Despacho.numero_despacho == numero_despacho
            ).first()
            
            if despacho:
                datos_actuales = despacho.datos_extraidos or {}
                datos_actuales[documento.tipo_documento] = documento.datos_extraidos
                despacho.datos_extraidos = datos_actuales
                despacho.fecha_actualizacion = datetime.now()
            
            db.commit()
            
            return {
                "message": "Documento procesado exitosamente",
                "documento_id": documento_id,
                "tipo": documento.tipo_documento,
                "datos_extraidos": documento.datos_extraidos
            }
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Error procesando documento: {response.text}"
            )
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando documento: {str(e)}")

@app.post("/despachos/{numero_despacho}/procesar")
async def procesar_despacho(
    numero_despacho: str,
    forzar: bool = False,
    db: Session = Depends(get_db),
    authorization: str = Header(None)
):
    """Procesar todos los documentos de un despacho"""
    despacho = db.query(Despacho).filter(
        Despacho.numero_despacho == numero_despacho
    ).first()
    
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    # Obtener documentos no procesados (o todos si forzar=True)
    if forzar:
        documentos = db.query(Documento).filter(
            Documento.numero_despacho == numero_despacho
        ).all()
    else:
        documentos = db.query(Documento).filter(
            Documento.numero_despacho == numero_despacho,
            Documento.procesado == False
        ).all()
    
    if not documentos:
        return {"message": "No hay documentos para procesar", "total_procesados": 0}
    
    despacho.estado = "procesando"
    db.commit()
    
    resultados = []
    datos_totales = {}
    
    for doc in documentos:
        try:
            if doc.contenido_base64:
                pdf_bytes = base64.b64decode(doc.contenido_base64)
                
                # Determinar endpoint según tipo
                endpoint = "invoice" if "factura" in doc.tipo_documento.lower() else "transport"
                
                from io import BytesIO
                files = {
                    'file': (doc.nombre_archivo, BytesIO(pdf_bytes), 'application/pdf')
                }
                
                data = {
                    'numero_despacho': numero_despacho,
                    'tipo_documento': doc.tipo_documento
                }
                
                headers = {}
                if authorization:
                    headers['Authorization'] = authorization
                
                response = requests.post(
                    f"http://api-docs:8002/process/{endpoint}",
                    files=files,
                    data=data,
                    headers=headers
                )
                
                if response.status_code == 200:
                    result = response.json()
                    
                    doc.datos_extraidos = result.get('extracted_data', {})
                    doc.procesado = True
                    doc.fecha_procesamiento = datetime.now()
                    
                    datos_totales[doc.tipo_documento] = doc.datos_extraidos
                    
                    resultados.append({
                        "documento_id": doc.id,
                        "tipo": doc.tipo_documento,
                        "estado": "procesado"
                    })
                else:
                    resultados.append({
                        "documento_id": doc.id,
                        "tipo": doc.tipo_documento,
                        "estado": "error",
                        "error": response.text
                    })
                    
        except Exception as e:
            resultados.append({
                "documento_id": doc.id,
                "tipo": doc.tipo_documento,
                "estado": "error",
                "error": str(e)
            })
    
    despacho.datos_extraidos = datos_totales
    despacho.estado = "completo"
    despacho.fecha_actualizacion = datetime.now()
    
    db.commit()
    
    return {
        "message": "Procesamiento completado",
        "resultados": resultados,
        "total_procesados": len([r for r in resultados if r["estado"] == "procesado"]),
        "total_errores": len([r for r in resultados if r["estado"] == "error"])
    }

@app.get("/despachos/{numero_despacho}/datos")
async def obtener_datos_despacho(
    numero_despacho: str,
    db: Session = Depends(get_db)
):
    """Obtener datos estructurados extraídos del despacho"""
    despacho = db.query(Despacho).filter(
        Despacho.numero_despacho == numero_despacho
    ).first()
    
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    return {
        "numero_despacho": numero_despacho,
        "estado": despacho.estado,
        "fecha_actualizacion": despacho.fecha_actualizacion.isoformat(),
        "datos_extraidos": despacho.datos_extraidos or {}
    }

@app.post("/despachos/{numero_despacho}/procedimiento")
async def crear_procedimiento(
    numero_despacho: str,
    procedimiento_data: ProcedimientoCreate,
    db: Session = Depends(get_db)
):
    """Crear un nuevo procedimiento para un despacho"""
    despacho = db.query(Despacho).filter(
        Despacho.numero_despacho == numero_despacho
    ).first()
    
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    nuevo_proc = Procedimiento(
        numero_despacho=numero_despacho,
        tipo_procedimiento=procedimiento_data.tipo_procedimiento,
        usuario_asignado=procedimiento_data.usuario_asignado,
        estado="pendiente",
        datos={}
    )
    
    db.add(nuevo_proc)
    db.commit()
    db.refresh(nuevo_proc)
    
    return {
        "message": "Procedimiento creado",
        "procedimiento_id": nuevo_proc.id,
        "tipo": nuevo_proc.tipo_procedimiento
    }

@app.put("/procedimiento/{procedimiento_id}/asignar/{usuario}")
async def asignar_procedimiento(
    procedimiento_id: int,
    usuario: str,
    db: Session = Depends(get_db)
):
    """Asignar un procedimiento a un usuario"""
    proc = db.query(Procedimiento).filter(
        Procedimiento.id == procedimiento_id
    ).first()
    
    if not proc:
        raise HTTPException(status_code=404, detail="Procedimiento no encontrado")
    
    proc.usuario_asignado = usuario
    proc.estado = "en_progreso"
    proc.fecha_inicio = datetime.now()
    
    db.commit()
    
    return {"message": f"Procedimiento asignado a {usuario}"}

@app.put("/procedimiento/{procedimiento_id}/completar")
async def completar_procedimiento(
    procedimiento_id: int,
    datos: Optional[Dict[str, Any]] = None,
    db: Session = Depends(get_db)
):
    """Marcar un procedimiento como completado"""
    proc = db.query(Procedimiento).filter(
        Procedimiento.id == procedimiento_id
    ).first()
    
    if not proc:
        raise HTTPException(status_code=404, detail="Procedimiento no encontrado")
    
    proc.estado = "completado"
    proc.fecha_fin = datetime.now()
    if datos:
        proc.datos = datos
    
    db.commit()
    
    return {"message": "Procedimiento completado"}

@app.get("/despachos")
async def listar_despachos(
    limit: int = 25,
    offset: int = 0,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Listar despachos con paginación y búsqueda"""
    query = db.query(Despacho)
    
    # Filtro de búsqueda si se proporciona
    if search:
        query = query.filter(
            Despacho.numero_despacho.ilike(f"%{search}%")
        )
    
    # Ordenar por fecha de actualización descendente (más recientes primero)
    query = query.order_by(Despacho.fecha_actualizacion.desc())
    
    # Obtener total para paginación
    total = query.count()
    
    # Aplicar paginación
    despachos = query.offset(offset).limit(limit).all()
    
    despachos_data = []
    for desp in despachos:
        documentos_requeridos = DOCUMENTOS_REQUERIDOS  # Siempre los mismos
        documentos_presentes = desp.documentos_presentes or []
        porcentaje = (len(documentos_presentes) / len(documentos_requeridos) * 100) if documentos_requeridos else 0
        
        despachos_data.append({
            "numero_despacho": desp.numero_despacho,
            "estado": desp.estado,
            "fecha_creacion": desp.fecha_creacion.isoformat(),
            "fecha_actualizacion": desp.fecha_actualizacion.isoformat(),
            "porcentaje_completitud": round(porcentaje, 1),
            "documentos_presentes": len(documentos_presentes),
            "documentos_requeridos": len(documentos_requeridos),
            "puede_procesar": len(documentos_presentes) == len(documentos_requeridos)
        })
    
    return {
        "despachos": despachos_data,
        "total": total,
        "limit": limit,
        "offset": offset,
        "total_pages": (total + limit - 1) // limit,
        "current_page": (offset // limit) + 1,
        "has_next": offset + limit < total,
        "has_prev": offset > 0
    }

@app.post("/despachos/{numero_despacho}/procesar")
async def procesar_documentos_despacho(
    numero_despacho: str,
    forzar: bool = Query(False),
    authorization: str = Header(None),
    db: Session = Depends(get_db)
):
    """Procesar documentos del despacho usando nuevo workflow"""
    despacho = db.query(Despacho).filter(Despacho.numero_despacho == numero_despacho).first()
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    if despacho.estado == "completo" and not forzar:
        raise HTTPException(status_code=400, detail="Despacho ya procesado. Use forzar=true para reprocesar")
    
    # Buscar documento principal (PDF multi-página)
    documento_principal = db.query(Documento).filter(
        Documento.numero_despacho == numero_despacho,
        Documento.tipo_documento == "documento_principal"
    ).first()
    
    if not documento_principal or not documento_principal.contenido_base64:
        raise HTTPException(status_code=400, detail="No se encontró documento principal para procesar")
    
    despacho.estado = "procesando"
    db.commit()
    
    try:
        # Procesar con nuevo workflow
        pdf_bytes = base64.b64decode(documento_principal.contenido_base64)
        
        files = {
            'file': (documento_principal.nombre_archivo, BytesIO(pdf_bytes), 'application/pdf')
        }
        
        data = {
            'numero_despacho': numero_despacho
        }
        
        headers = {}
        if authorization:
            headers['Authorization'] = authorization
        
        # Llamar al nuevo endpoint de procesamiento
        response = requests.post(
            f"http://api-docs:8002/process/dispatch",
            files=files,
            data=data,
            headers=headers,
            timeout=600  # 10 minutos para procesamiento completo
        )
        
        if response.status_code == 200:
            result = response.json()
            
            # El nuevo endpoint ya guarda los documentos individualmente
            despacho.estado = "completo"
            despacho.fecha_actualizacion = datetime.now()
            db.commit()
            
            return {
                "message": "Procesamiento completado con nuevo workflow",
                "documentos_identificados": result['result']['documentos_procesados'],
                "resultado": result['result']
            }
        else:
            despacho.estado = "error"
            db.commit()
            raise HTTPException(status_code=500, detail=f"Error en procesamiento: {response.text}")
            
    except Exception as e:
        despacho.estado = "error"
        db.commit()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/despachos/{numero_despacho}/upload-principal")
async def upload_documento_principal(
    numero_despacho: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Subir documento principal multi-página para procesamiento"""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF")
    
    # Verificar que el despacho existe
    despacho = db.query(Despacho).filter(Despacho.numero_despacho == numero_despacho).first()
    if not despacho:
        raise HTTPException(status_code=404, detail="Despacho no encontrado")
    
    try:
        # Leer contenido
        contents = await file.read()
        
        if not contents.startswith(b'%PDF'):
            raise HTTPException(status_code=400, detail="El archivo no es un PDF válido")
        
        # Convertir a base64
        contenido_b64 = base64.b64encode(contents).decode('utf-8')
        
        # Buscar si ya existe documento principal
        doc_existente = db.query(Documento).filter(
            Documento.numero_despacho == numero_despacho,
            Documento.tipo_documento == "documento_principal"
        ).first()
        
        if doc_existente:
            # Actualizar existente
            doc_existente.nombre_archivo = file.filename
            doc_existente.contenido_base64 = contenido_b64
            doc_existente.fecha_carga = datetime.now()
            doc_existente.procesado = False
        else:
            # Crear nuevo
            nuevo_doc = Documento(
                numero_despacho=numero_despacho,
                tipo_documento="documento_principal",
                nombre_archivo=file.filename,
                contenido_base64=contenido_b64,
                procesado=False,
                fecha_carga=datetime.now()
            )
            db.add(nuevo_doc)
        
        # Resetear estado del despacho
        despacho.estado = "pendiente"
        despacho.fecha_actualizacion = datetime.now()
        
        db.commit()
        
        return {
            "message": "Documento principal cargado correctamente",
            "filename": file.filename,
            "puede_procesar": True
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error cargando archivo: {str(e)}")

from io import BytesIO

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8003)