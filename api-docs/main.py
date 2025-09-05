# api-docs/main.py
from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Header, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, Dict, List, Any
from datetime import datetime
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import os
import base64
import json
import io
import re
import uuid
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configuración Azure
ENDPOINT = os.getenv('AZURE_FORM_RECOGNIZER_ENDPOINT', '')
API_KEY = os.getenv('AZURE_FORM_RECOGNIZER_KEY', '')

# IMPORTANTE: Los nombres de los modelos deben coincidir exactamente con los de Azure
# Si hay un typo en el .env (inovice_01 en lugar de invoice_01), corregirlo
DOCTYPE_MODEL_ID = os.getenv('DOCTYPE_MODEL_ID', 'doctype_01')
INVOICE_MODEL_ID = os.getenv('INVOICE_MODEL_ID', 'invoice_01')  # Verificar: era "inovice_01" en el .env
TRANSPORT_MODEL_ID = os.getenv('TRANSPORT_MODEL_ID', 'transport_01')

# Cliente Azure
document_analysis_client = DocumentAnalysisClient(
    endpoint=ENDPOINT,
    credential=AzureKeyCredential(API_KEY)
) if ENDPOINT and API_KEY else None

app = FastAPI(title="Document Processing API")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Storage en memoria para procesos
app.state.processed_dispatches = {}
app.state.processed_documents = {}

# ==================== FUNCIONES AUXILIARES ====================

def extract_invoice_data(result) -> Dict:
    """Extraer datos estructurados de una factura"""
    invoice_data = {
        "metadata": {
            "pages": len(result.pages) if hasattr(result, 'pages') else 0,
            "model_id": result.model_id if hasattr(result, 'model_id') else 'unknown',
            "processed_at": datetime.now().isoformat()
        },
        "vendor_information": {},
        "customer_information": {},
        "invoice_details": {},
        "totals": {},
        "line_items": [],
        "all_fields": {}
    }
    
    # Extraer campos del modelo custom
    if hasattr(result, 'documents') and result.documents:
        for doc in result.documents:
            if hasattr(doc, 'fields'):
                for field_name, field_value in doc.fields.items():
                    if hasattr(field_value, 'value') and field_value.value:
                        value = str(field_value.value)
                        invoice_data["all_fields"][field_name] = value
                        
                        # Clasificar por tipo de campo
                        field_lower = field_name.lower()
                        if 'vendor' in field_lower or 'supplier' in field_lower:
                            invoice_data["vendor_information"][field_name] = value
                        elif 'customer' in field_lower or 'buyer' in field_lower:
                            invoice_data["customer_information"][field_name] = value
                        elif 'total' in field_lower or 'amount' in field_lower:
                            invoice_data["totals"][field_name] = value
                        elif 'item' in field_lower or 'line' in field_lower:
                            invoice_data["line_items"].append({field_name: value})
                        elif any(term in field_lower for term in ['invoice', 'number', 'date', 'due']):
                            invoice_data["invoice_details"][field_name] = value
    
    # Extraer key-value pairs si no hay campos del modelo custom
    if not invoice_data["all_fields"] and hasattr(result, 'key_value_pairs'):
        for kv_pair in result.key_value_pairs:
            if kv_pair.key and kv_pair.value:
                key = kv_pair.key.content if hasattr(kv_pair.key, 'content') else str(kv_pair.key)
                value = kv_pair.value.content if hasattr(kv_pair.value, 'content') else str(kv_pair.value)
                
                invoice_data["all_fields"][key] = value
                key_lower = key.lower()
                
                if any(term in key_lower for term in ['vendor', 'supplier', 'seller']):
                    invoice_data["vendor_information"][key] = value
                elif any(term in key_lower for term in ['customer', 'buyer', 'bill to']):
                    invoice_data["customer_information"][key] = value
                elif any(term in key_lower for term in ['total', 'amount', 'subtotal', 'tax']):
                    invoice_data["totals"][key] = value
                elif any(term in key_lower for term in ['invoice', 'number', 'date']):
                    invoice_data["invoice_details"][key] = value
    
    return invoice_data

def extract_transport_data(result) -> Dict:
    """Extraer datos estructurados de un documento de transporte"""
    transport_data = {
        "metadata": {
            "pages": len(result.pages) if hasattr(result, 'pages') else 0,
            "model_id": result.model_id if hasattr(result, 'model_id') else 'unknown',
            "processed_at": datetime.now().isoformat()
        },
        "shipper": {},
        "consignee": {},
        "transport_details": {},
        "goods": [],
        "all_fields": {}
    }
    
    # Extraer campos del modelo custom
    if hasattr(result, 'documents') and result.documents:
        for doc in result.documents:
            if hasattr(doc, 'fields'):
                for field_name, field_value in doc.fields.items():
                    if hasattr(field_value, 'value') and field_value.value:
                        value = str(field_value.value)
                        transport_data["all_fields"][field_name] = value
                        
                        field_lower = field_name.lower()
                        if 'shipper' in field_lower or 'sender' in field_lower:
                            transport_data["shipper"][field_name] = value
                        elif 'consignee' in field_lower or 'receiver' in field_lower:
                            transport_data["consignee"][field_name] = value
                        elif 'goods' in field_lower or 'cargo' in field_lower:
                            transport_data["goods"].append({field_name: value})
                        else:
                            transport_data["transport_details"][field_name] = value
    
    # Extraer key-value pairs si no hay campos del modelo custom
    if not transport_data["all_fields"] and hasattr(result, 'key_value_pairs'):
        for kv_pair in result.key_value_pairs:
            if kv_pair.key and kv_pair.value:
                key = kv_pair.key.content if hasattr(kv_pair.key, 'content') else str(kv_pair.key)
                value = kv_pair.value.content if hasattr(kv_pair.value, 'content') else str(kv_pair.value)
                
                transport_data["all_fields"][key] = value
                key_lower = key.lower()
                
                if any(term in key_lower for term in ['shipper', 'sender', 'exportador']):
                    transport_data["shipper"][key] = value
                elif any(term in key_lower for term in ['consignee', 'receiver', 'importador']):
                    transport_data["consignee"][key] = value
                elif any(term in key_lower for term in ['transport', 'vessel', 'container', 'booking']):
                    transport_data["transport_details"][key] = value
    
    return transport_data

def create_excel_from_dispatch(data: Dict) -> io.BytesIO:
    """Crear Excel con datos del despacho procesado"""
    wb = Workbook()
    
    # Hoja resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Encabezado
    ws_resumen['A1'] = "RESUMEN DEL DESPACHO"
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['A3'] = "Número Despacho:"
    ws_resumen['B3'] = data.get('numero_despacho', '')
    ws_resumen['A4'] = "Total Páginas:"
    ws_resumen['B4'] = data.get('total_paginas', 0)
    ws_resumen['A5'] = "Total Documentos:"
    ws_resumen['B5'] = data.get('total_documentos', 0)
    
    # Resumen por tipo
    ws_resumen['A7'] = "DOCUMENTOS POR TIPO"
    ws_resumen['A7'].font = Font(bold=True)
    ws_resumen['A7'].fill = header_fill
    ws_resumen['A7'].font = header_font
    ws_resumen.merge_cells('A7:B7')
    
    row = 8
    for tipo, cantidad in data.get('resumen', {}).items():
        ws_resumen[f'A{row}'] = tipo.replace('_', ' ').title()
        ws_resumen[f'B{row}'] = cantidad
        row += 1
    
    # Crear hoja para cada documento procesado
    for doc in data.get('documentos', []):
        # Crear hoja
        sheet_name = f"{doc['tipo'][:10]}_{doc['id'][-5:]}"
        ws_doc = wb.create_sheet(title=sheet_name)
        
        # Información del documento
        ws_doc['A1'] = f"DOCUMENTO: {doc['tipo'].upper()}"
        ws_doc['A1'].font = Font(bold=True, size=12)
        ws_doc['A3'] = "ID:"
        ws_doc['B3'] = doc['id']
        ws_doc['A4'] = "Páginas:"
        ws_doc['B4'] = doc['paginas']
        ws_doc['A5'] = "Procesado:"
        ws_doc['B5'] = "Sí" if doc['procesado'] else "No"
        
        # Datos extraídos
        if doc.get('datos_extraidos') and not doc['datos_extraidos'].get('error'):
            ws_doc['A7'] = "DATOS EXTRAÍDOS"
            ws_doc['A7'].font = Font(bold=True)
            ws_doc['A7'].fill = header_fill
            ws_doc['A7'].font = header_font
            ws_doc.merge_cells('A7:B7')
            
            row = 8
            datos = doc['datos_extraidos']
            
            # Facturas
            if doc['tipo'] == 'factura':
                # Vendedor
                if datos.get('vendor_information'):
                    ws_doc[f'A{row}'] = "VENDEDOR"
                    ws_doc[f'A{row}'].font = Font(bold=True)
                    row += 1
                    for k, v in datos['vendor_information'].items():
                        ws_doc[f'A{row}'] = k.replace('_', ' ').title()
                        ws_doc[f'B{row}'] = str(v)
                        row += 1
                    row += 1
                
                # Cliente
                if datos.get('customer_information'):
                    ws_doc[f'A{row}'] = "CLIENTE"
                    ws_doc[f'A{row}'].font = Font(bold=True)
                    row += 1
                    for k, v in datos['customer_information'].items():
                        ws_doc[f'A{row}'] = k.replace('_', ' ').title()
                        ws_doc[f'B{row}'] = str(v)
                        row += 1
                    row += 1
                
                # Totales
                if datos.get('totals'):
                    ws_doc[f'A{row}'] = "TOTALES"
                    ws_doc[f'A{row}'].font = Font(bold=True)
                    row += 1
                    for k, v in datos['totals'].items():
                        ws_doc[f'A{row}'] = k.replace('_', ' ').title()
                        ws_doc[f'B{row}'] = str(v)
                        row += 1
            
            # Transportes
            elif doc['tipo'] == 'transporte':
                # Remitente
                if datos.get('shipper'):
                    ws_doc[f'A{row}'] = "REMITENTE"
                    ws_doc[f'A{row}'].font = Font(bold=True)
                    row += 1
                    for k, v in datos['shipper'].items():
                        ws_doc[f'A{row}'] = k.replace('_', ' ').title()
                        ws_doc[f'B{row}'] = str(v)
                        row += 1
                    row += 1
                
                # Destinatario
                if datos.get('consignee'):
                    ws_doc[f'A{row}'] = "DESTINATARIO"
                    ws_doc[f'A{row}'].font = Font(bold=True)
                    row += 1
                    for k, v in datos['consignee'].items():
                        ws_doc[f'A{row}'] = k.replace('_', ' ').title()
                        ws_doc[f'B{row}'] = str(v)
                        row += 1
    
    # Ajustar anchos de columna
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

# ==================== ENDPOINTS ====================

@app.get("/")
async def root():
    return {
        "service": "Document Processing API",
        "status": "running",
        "models": {
            "doctype": DOCTYPE_MODEL_ID,
            "invoice": INVOICE_MODEL_ID,
            "transport": TRANSPORT_MODEL_ID
        }
    }

@app.get("/health")
async def health():
    return {
        "status": "healthy" if document_analysis_client else "degraded",
        "service": "api-docs",
        "azure_configured": bool(document_analysis_client),
        "timestamp": datetime.now().isoformat(),
        "models": {
            "doctype": DOCTYPE_MODEL_ID,
            "invoice": INVOICE_MODEL_ID,
            "transport": TRANSPORT_MODEL_ID
        }
    }

@app.post("/process/automatic")
async def process_automatic(
    file: UploadFile = File(...),
    numero_despacho: str = Form(...)
):
    """Procesar documento con identificación automática y separación"""
    if not document_analysis_client:
        raise HTTPException(status_code=500, detail="Azure Document Intelligence no configurado")
    
    process_id = str(uuid.uuid4())
    
    try:
        # Leer archivo
        contents = await file.read()
        
        if not contents.startswith(b'%PDF'):
            raise HTTPException(status_code=400, detail="El archivo no es un PDF válido")
        
        # Ejecutar workflow completo
        from document_processor import process_dispatch_workflow
        resultado = process_dispatch_workflow(contents, numero_despacho)
        
        if resultado.get("error"):
            raise HTTPException(status_code=500, detail=resultado["error"])
        
        # Guardar para descarga posterior
        app.state.processed_dispatches[process_id] = resultado
        
        return {
            "id": process_id,
            "numero_despacho": numero_despacho,
            "filename": file.filename,
            "resultado": resultado,
            "json_url": f"/download/{process_id}/json",
            "excel_url": f"/download/{process_id}/excel"
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/process/invoice")
async def process_invoice(
    file: UploadFile = File(...),
    numero_despacho: Optional[str] = Form(None)
):
    """Procesar una factura individual usando el modelo invoice_01"""
    if not document_analysis_client:
        raise HTTPException(status_code=500, detail="Azure Document Intelligence no configurado")
    
    process_id = str(uuid.uuid4())
    
    try:
        contents = await file.read()
        
        if not contents.startswith(b'%PDF'):
            raise HTTPException(status_code=400, detail="El archivo no es un PDF válido")
        
        # Analizar con Azure
        poller = document_analysis_client.begin_analyze_document(
            INVOICE_MODEL_ID,
            document=contents
        )
        result = poller.result()
        
        # Extraer datos
        invoice_data = extract_invoice_data(result)
        
        # Guardar para descarga
        app.state.processed_documents[process_id] = {
            'data': invoice_data,
            'type': 'invoice',
            'filename': file.filename,
            'numero_despacho': numero_despacho
        }
        
        return {
            "id": process_id,
            "filename": file.filename,
            "document_type": "invoice",
            "extracted_data": invoice_data,
            "json_url": f"/download/doc/{process_id}/json"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/process/transport")
async def process_transport(
    file: UploadFile = File(...),
    numero_despacho: Optional[str] = Form(None)
):
    """Procesar un documento de transporte individual usando el modelo transport_01"""
    if not document_analysis_client:
        raise HTTPException(status_code=500, detail="Azure Document Intelligence no configurado")
    
    process_id = str(uuid.uuid4())
    
    try:
        contents = await file.read()
        
        if not contents.startswith(b'%PDF'):
            raise HTTPException(status_code=400, detail="El archivo no es un PDF válido")
        
        # Analizar con Azure
        poller = document_analysis_client.begin_analyze_document(
            TRANSPORT_MODEL_ID,
            document=contents
        )
        result = poller.result()
        
        # Extraer datos
        transport_data = extract_transport_data(result)
        
        # Guardar para descarga
        app.state.processed_documents[process_id] = {
            'data': transport_data,
            'type': 'transport',
            'filename': file.filename,
            'numero_despacho': numero_despacho
        }
        
        return {
            "id": process_id,
            "filename": file.filename,
            "document_type": "transport",
            "extracted_data": transport_data,
            "json_url": f"/download/doc/{process_id}/json"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download/{process_id}/json")
async def download_dispatch_json(process_id: str):
    """Descargar datos de despacho procesado como JSON"""
    if process_id not in app.state.processed_dispatches:
        raise HTTPException(status_code=404, detail="Proceso no encontrado")
    
    data = app.state.processed_dispatches[process_id]
    
    # Crear JSON sin los PDFs base64 para reducir tamaño
    export_data = {
        "numero_despacho": data['numero_despacho'],
        "timestamp": data['timestamp'],
        "total_paginas": data['total_paginas'],
        "total_documentos": data['total_documentos'],
        "resumen": data['resumen'],
        "documentos_procesados": []
    }
    
    for doc in data['documentos']:
        doc_export = {
            "id": doc['id'],
            "tipo": doc['tipo'],
            "paginas": doc['paginas'],
            "procesado": doc['procesado'],
            "datos_extraidos": doc['datos_extraidos']
        }
        export_data["documentos_procesados"].append(doc_export)
    
    json_content = json.dumps(export_data, indent=2, ensure_ascii=False)
    
    return Response(
        content=json_content,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=despacho_{data['numero_despacho']}.json"
        }
    )

@app.get("/download/{process_id}/excel")
async def download_dispatch_excel(process_id: str):
    """Descargar datos de despacho procesado como Excel"""
    if process_id not in app.state.processed_dispatches:
        raise HTTPException(status_code=404, detail="Proceso no encontrado")
    
    data = app.state.processed_dispatches[process_id]
    excel_file = create_excel_from_dispatch(data)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=despacho_{data['numero_despacho']}.xlsx"
        }
    )

@app.get("/download/doc/{process_id}/json")
async def download_document_json(process_id: str):
    """Descargar datos de documento individual como JSON"""
    if process_id not in app.state.processed_documents:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    doc_data = app.state.processed_documents[process_id]
    json_content = json.dumps(doc_data['data'], indent=2, ensure_ascii=False)
    
    return Response(
        content=json_content,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename={doc_data['type']}_{process_id}.json"
        }
    )

@app.post("/debug/test-model")
async def test_model(
    file: UploadFile = File(...),
    model_type: str = Form(...)  # "doctype", "invoice", o "transport"
):
    """Endpoint de debug para probar modelos específicos"""
    if not document_analysis_client:
        raise HTTPException(status_code=500, detail="Azure no configurado")
    
    try:
        contents = await file.read()
        
        # Seleccionar modelo
        model_map = {
            "doctype": DOCTYPE_MODEL_ID,
            "invoice": INVOICE_MODEL_ID,
            "transport": TRANSPORT_MODEL_ID
        }
        
        model_id = model_map.get(model_type)
        if not model_id:
            raise HTTPException(status_code=400, detail="Tipo de modelo inválido")
        
        # Procesar
        poller = document_analysis_client.begin_analyze_document(
            model_id,
            document=contents
        )
        result = poller.result()
        
        # Información de debug
        debug_info = {
            "model_id_used": model_id,
            "pages_count": len(result.pages) if hasattr(result, 'pages') else 0,
            "has_documents": hasattr(result, 'documents') and bool(result.documents),
            "has_key_value_pairs": hasattr(result, 'key_value_pairs') and bool(result.key_value_pairs),
        }
        
        # Campos disponibles
        if hasattr(result, 'documents') and result.documents:
            debug_info["document_fields"] = list(result.documents[0].fields.keys()) if hasattr(result.documents[0], 'fields') else []
            
            # Para doctype, mostrar el tipo detectado
            if model_type == "doctype" and hasattr(result.documents[0], 'doc_type'):
                debug_info["detected_type"] = result.documents[0].doc_type
        
        # Extraer datos según tipo
        extracted_data = {}
        if model_type == "invoice":
            extracted_data = extract_invoice_data(result)
        elif model_type == "transport":
            extracted_data = extract_transport_data(result)
        elif model_type == "doctype" and hasattr(result, 'documents'):
            # Para doctype, extraer información de clasificación
            for doc in result.documents:
                if hasattr(doc, 'doc_type'):
                    extracted_data["document_type"] = doc.doc_type
                if hasattr(doc, 'confidence'):
                    extracted_data["confidence"] = doc.confidence
        
        return {
            "debug_info": debug_info,
            "extracted_data": extracted_data
        }
        
    except Exception as e:
        import traceback
        return {
            "error": str(e),
            "traceback": traceback.format_exc()
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8002)